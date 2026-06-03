import pandas as pd
import os
import sys
import glob
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# --- COLOR CONSTANTS ---
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
DARK_GREEN_FILL = PatternFill(start_color='375623', end_color='375623', fill_type='solid')
DARK_RED_FILL = PatternFill(start_color='9C0006', end_color='9C0006', fill_type='solid')
NEUTRAL_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
WARNING_FILL = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
WHITE_FONT = Font(color='FFFFFF')

def get_style(col_name, val):
    """Returns (fill, font) based on the specific column rules."""
    try:
        # ── BOOLEAN COLUMNS ──
        if col_name in ['g200', 'g050', 'g020', 'bbup', 'bbsq', 'vspk']:
            if str(val).strip().lower() in ['true', '1', '1.0']: return GREEN_FILL, None
            if str(val).strip().lower() in ['false', '0', '0.0']: return RED_FILL, None

        # ── DIRECTIONAL COLUMNS ──
        if col_name == 'chan':
            val = float(val)
            if val > 0: return GREEN_FILL, None
            if val < 0: return RED_FILL, None
            if val == 0: return NEUTRAL_FILL, None
        
        if col_name == 'CMF_20':
            val = float(val)
            if val > 0.05: return GREEN_FILL, None
            if val < -0.05: return RED_FILL, None
            return NEUTRAL_FILL, None
            
        if col_name == 'SUPERTd_7_3.0':
            if val == 1.0: return GREEN_FILL, None
            if val == -1.0: return RED_FILL, None
            return NEUTRAL_FILL, None
            
        if col_name == 'SQZ_ON' and val == 1:
            return GREEN_FILL, None
            
        if col_name == 'EFI_13':
            val = float(val)
            if val > 0: return GREEN_FILL, None
            if val < 0: return RED_FILL, None
            if val == 0: return NEUTRAL_FILL, None
            
        if col_name == 'tren':
            if val == "Uptrend": return GREEN_FILL, None
            if val == "Downtrend": return RED_FILL, None
            if val == "Sideways": return NEUTRAL_FILL, None
            
        if col_name == 'tstr':
            if val == "Strong": return GREEN_FILL, None
            if val == "Weak": return RED_FILL, None
            if val == "Moderate": return NEUTRAL_FILL, None
            
        if col_name == 'vtrd':
            if val == "Increasing": return GREEN_FILL, None
            if val == "Decreasing": return RED_FILL, None
            
        if col_name == 'stge':
            if "Stage 2" in str(val): return GREEN_FILL, None
            if "Stage 4" in str(val): return RED_FILL, None
            if any(s in str(val) for s in ["Stage 1", "Stage 3"]): return NEUTRAL_FILL, None

        # ── RANGE COLUMNS ──
        if col_name == 'rsi':
            if val >= 80: return DARK_RED_FILL, WHITE_FONT
            if val >= 70: return WARNING_FILL, None
            if val >= 55: return DARK_GREEN_FILL, WHITE_FONT
            if val >= 45: return GREEN_FILL, None
            if val >= 35: return NEUTRAL_FILL, None
            return RED_FILL, None
            
        if col_name == 'vola':
            if val >= 80: return DARK_RED_FILL, WHITE_FONT
            if val >= 55: return RED_FILL, None
            if val >= 35: return WARNING_FILL, None
            if val >= 20: return GREEN_FILL, None
            return DARK_GREEN_FILL, WHITE_FONT
            
        if col_name == 'bbbw':
            if val >= 0.50: return DARK_RED_FILL, WHITE_FONT
            if val >= 0.35: return RED_FILL, None
            if val >= 0.20: return WARNING_FILL, None
            if val >= 0.10: return GREEN_FILL, None
            return DARK_GREEN_FILL, WHITE_FONT
            
        if col_name == 'STOCHk_14_3_3':
            if val >= 85: return DARK_RED_FILL, WHITE_FONT
            if val >= 70: return RED_FILL, None
            if val >= 45: return WARNING_FILL, None
            if val >= 25: return GREEN_FILL, None
            return DARK_GREEN_FILL, WHITE_FONT
            
        if col_name == 'wrsi':
            if val > 75: return WARNING_FILL, None
            if val >= 50: return GREEN_FILL, None
            if val >= 45: return NEUTRAL_FILL, None
            return RED_FILL, None
            
        if col_name == 'adx':
            if val >= 25: return DARK_GREEN_FILL, WHITE_FONT # Strong trend
            if val >= 20: return GREEN_FILL, None
            if val >= 15: return NEUTRAL_FILL, None
            return RED_FILL, None
            
        if col_name == 'rvol':
            if val >= 3.0: return DARK_GREEN_FILL, WHITE_FONT # Institutional surge
            if val >= 1.5: return GREEN_FILL, None
            if val >= 0.8: return NEUTRAL_FILL, None
            return RED_FILL, None
            
        if col_name == 'WILLR_14':
            if val < -80: return GREEN_FILL, None
            if val > -20: return RED_FILL, None
            
        if col_name == 'RSI_2':
            if val <= 10: return GREEN_FILL, None # Oversold bounce potential
            if val >= 90: return WARNING_FILL, None
            
        if col_name == 'DlPer':
            if val >= 65: return DARK_GREEN_FILL, WHITE_FONT # Heavy accumulation
            if val >= 45: return GREEN_FILL, None
            if val >= 20: return NEUTRAL_FILL, None
            return RED_FILL, None
            
        if col_name == 'delt':
            if val >= 30: return DARK_GREEN_FILL, WHITE_FONT # Significant room to grow
            if val >= 15: return GREEN_FILL, None
            if val >= 7: return NEUTRAL_FILL, None
            return RED_FILL, None # Near 52W High (Expensive)
            
        if col_name == 'zone':
            if val == "Premium": return DARK_RED_FILL, WHITE_FONT
            if val == "Near Premium": return RED_FILL, None
            if val == "Equilibrium": return NEUTRAL_FILL, None
            if val == "Near Discount": return DARK_GREEN_FILL, WHITE_FONT
            if val == "Discount": return GREEN_FILL, None

        # ── VOLUME & ACTIVITY RANKS ──
        if col_name in ['vrnk', 'rrnk', 'arnk']:
            if val <= 50: return DARK_GREEN_FILL, WHITE_FONT # Top Tier Activity
            if val <= 150: return GREEN_FILL, None
            if val >= 400: return RED_FILL, None

    except:
        pass
    return None, None

def colorize_snapshot(csv_path):
    if not os.path.exists(csv_path):
        print(f"Error: File {csv_path} not found.")
        return

    xlsx_path = csv_path.replace('.csv', '.xlsx')
    print(f"Processing {csv_path}...")

    # Load Data
    df = pd.read_csv(csv_path)
    
    # Split Patterns to second sheet
    pattern_cols = ['psta', 'pend', 'ppnt']
    # Check if cols exist before splitting
    actual_pattern_cols = [c for c in pattern_cols if c in df.columns]
    
    df_main = df.drop(columns=actual_pattern_cols)
    df_patterns = df[['symb'] + actual_pattern_cols].copy() if actual_pattern_cols else pd.DataFrame()

    # Create Excel Writer
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df_main.to_excel(writer, sheet_name='Main Analysis', index=False)
        if not df_patterns.empty:
            df_patterns.to_excel(writer, sheet_name='Chart Patterns', index=False)
        
        # Legend Sheet Data
        legend_data = [
            ["Column", "Condition", "Color", "Meaning"],
            ["Booleans", "True", "Light Green", "Bullish / Signal Active"],
            ["Booleans", "False", "Light Red", "Bearish / Signal Inactive"],
            ["chan", "> 0", "Light Green", "Price Gain"],
            ["CMF_20", "> 0.05", "Light Green", "Money Inflow"],
            ["SUPERTd", "1.0", "Light Green", "Uptrend Confirm"],
            ["SQZ_ON", "1", "Light Green", "Coiling / Squeeze"],
            ["tren", "Uptrend", "Light Green", "Bullish Trend"],
            ["rsi", "Intensity Gradient", "G -> Y -> R", "Momentum strength vs Overbought risk"],
            ["vola", "Intensity Gradient", "G -> Y -> R", "Stability vs Annualized Volatility Risk"],
            ["bbbw", "Intensity Gradient", "G -> Y -> R", "Tight Squeeze to Extreme Expansion"],
            ["STOCHk", "Intensity Gradient", "G -> Y -> R", "Oversold Accumulation to Overbought Pivot"],
            ["adx", ">= 25", "Light Green", "Strong Trend"],
            ["rvol", ">= 3.0", "Dark Green", "Surge: 3x normal volume"],
            ["vrnk/arnk", "<= 50", "Dark Green", "Market Liquidity Leader"],
            ["delt", ">= 30%", "Dark Green / White Font", "Significant room to grow"],
            ["delt", "<= 7%", "Light Red", "Near 52W High (Mean Reversion Risk)"],
            ["zone", "Premium", "Dark Red", "Overbought Highs (Mean Reversion Risk)"],
            ["zone", "Near Discount", "Dark Green", "Accumulation Zone (Near Support)"],
            ["zone", "Equilibrium", "Yellow", "Mid-range Consolidation"]
        ]
        pd.DataFrame(legend_data).to_excel(writer, sheet_name='Legend', index=False, header=False)

    # Re-open with openpyxl for styling
    wb = load_workbook(xlsx_path)
    ws = wb['Main Analysis']

    # 1. Freeze Top Row
    ws.freeze_panes = 'A2'

    # 2. Apply Colors and Auto-Fit
    headers = [cell.value for cell in ws[1]]
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for col_idx, cell in enumerate(row):
            col_name = headers[col_idx]
            
            # Apply readable number formatting to raw volume
            if col_name == 'volu':
                cell.number_format = '#,##0'

            fill, font = get_style(col_name, cell.value)
            if fill: cell.fill = fill
            if font: cell.font = font

    # Auto-fit Column Widths
    for i, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(i)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 40) # Cap width

    # Style Legend Sheet
    if 'Legend' in wb.sheetnames:
        lws = wb['Legend']
        for row in lws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left')
        lws.column_dimensions['A'].width = 20
        lws.column_dimensions['B'].width = 30
        lws.column_dimensions['C'].width = 25
        lws.column_dimensions['D'].width = 40

    # 3. Convert to Excel Tables for Filtering
    for sheet_name, table_name in [('Main Analysis', 'AnalysisTable'), ('Chart Patterns', 'PatternsTable')]:
        if sheet_name in wb.sheetnames:
            target_ws = wb[sheet_name]
            if target_ws.max_row > 1:
                # Define the table range (e.g., A1:Z100)
                last_col = get_column_letter(target_ws.max_column)
                tab_range = f"A1:{last_col}{target_ws.max_row}"
                
                tab = Table(displayName=table_name, ref=tab_range)
                # Use a light style so it doesn't clash with our conditional colors
                style = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False,
                                       showLastColumn=False, showRowStripes=False, showColumnStripes=False)
                tab.tableStyleInfo = style
                target_ws.add_table(tab)

    wb.save(xlsx_path)
    print(f"Done! Excel file saved: {xlsx_path}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        target = sys.argv[1]
    else:
        # Sort by OS modification time to ensure we actually get the most recently created file
        snapshot_files = sorted(glob.glob("*snapshot.csv"), key=os.path.getmtime)
        snapshot_all_files = sorted(glob.glob("*snapshot_all.csv"), key=os.path.getmtime)
        # Filter out 'all' snapshots from the standard list to avoid duplicates
        snapshot_files = [f for f in snapshot_files if not f.endswith("_all.csv")]

        if snapshot_files and snapshot_all_files:
            print("Choose snapshot source:")
            print("1. snapshot.csv")
            print("2. snapshot_all.csv")
            choice = input("Enter choice [default 1]: ").strip()
            if choice == '2':
                target = snapshot_all_files[-1]
            else:
                target = snapshot_files[-1]
        elif snapshot_files:
            target = snapshot_files[-1]
        elif snapshot_all_files:
            target = snapshot_all_files[-1]
        else:
            print("No snapshot CSV found in current directory.")
            sys.exit(1)
    colorize_snapshot(target)
